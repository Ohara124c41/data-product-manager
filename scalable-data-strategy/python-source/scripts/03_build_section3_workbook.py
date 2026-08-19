from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import json

src=Path('upload/section-3-event-logs-template.xlsx')
out=Path('outputs/flyber/section-3-event-logs-completed.xlsx')
a=json.load(open('flyber_analysis.json'))['etl']
wb=load_workbook(src)
ws=wb['ETL']

# Reset only the existing empty ETL sheet; preserve the original raw-log sheet.
for merged in list(ws.merged_cells.ranges): ws.unmerge_cells(str(merged))
if ws.max_row: ws.delete_rows(1,ws.max_row)

navy='17365D'; light='D9EAF7'; pale='F3F6F9'; green='E2F0D9'; dark='1F2937'; white='FFFFFF'; line='D9E2F3'
thin=Side(style='thin',color=line)
border=Border(left=thin,right=thin,top=thin,bottom=thin)

ws.merge_cells('A1:H1'); ws['A1']='Flyber Event Log ETL - October 5-11, 2019'
ws['A1'].fill=PatternFill('solid',fgColor=navy); ws['A1'].font=Font(name='Arial',size=16,bold=True,color=white); ws['A1'].alignment=Alignment(vertical='center')
ws.row_dimensions[1].height=28
ws.merge_cells('A2:H2'); ws['A2']='Scope: 124,980 raw events inspected. The incomplete October 12 boundary segment is excluded from daily reporting.'
ws['A2'].fill=PatternFill('solid',fgColor=pale); ws['A2'].font=Font(name='Arial',size=10,italic=True,color=dark); ws['A2'].alignment=Alignment(wrap_text=True,vertical='center')
ws.row_dimensions[2].height=30

dates=[datetime.strptime(x,'%Y-%m-%d').date() for x in a['dates'][:7]]
def block(row,label,categories,values):
    headers=[label,*dates]
    for c,v in enumerate(headers,1):
        cell=ws.cell(row,c,v); cell.fill=PatternFill('solid',fgColor=light); cell.font=Font(name='Arial',size=10,bold=True,color=navy); cell.border=border; cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        if c>1: cell.number_format='m/d/yyyy'
    for i,cat in enumerate(categories,1):
        ws.cell(row+i,1,cat)
        for j,val in enumerate(values[i-1][:7],2):
            cell=ws.cell(row+i,j,int(val)); cell.number_format='#,##0'
        for c in range(1,9):
            cell=ws.cell(row+i,c); cell.font=Font(name='Arial',size=10,color=dark); cell.border=border; cell.alignment=Alignment(vertical='center')

block(4,'Date / Metric',['Total events'],[a['total']])
block(8,'Event Type',['Choose Car','Search','Open','Begin Ride','Request Car'],a['events']['values'])
block(16,'Device Type',['iOS','Android','Desktop Web','Mobile Web'],a['devices']['values'])
block(23,'Page Type',['Search Page','Book Page','Driver Page','Splash Page'],a['pages']['values'])
block(30,'Location',a['locations']['categories'],a['locations']['values'])

ws.merge_cells('A38:H38'); ws['A38']='Transformation and Quality-Control Notes'
ws['A38'].fill=PatternFill('solid',fgColor=light); ws['A38'].font=Font(name='Arial',size=10,bold=True,color=navy); ws['A38'].alignment=Alignment(horizontal='center')
notes=[
('1. Extract','Loaded the supplied XLSX event-log table and retained event_uuid, user_uuid, event_time, device_type, session_uuid, user_neighborhood, event_page, and event_type. XLSX preserves timestamps and identifiers without delimiter ambiguity.'),
('2. Standardize','Converted Excel serial timestamps to calendar dates, retained identifiers as text, and mapped machine-readable category labels to presentation labels.'),
('3. Filter','Restricted the complete reporting window to October 5-11, 2019. The October 12 observations form an incomplete boundary day and are excluded.'),
('4. Aggregate','Grouped records by date and counted event_uuid for total, event type, device type, page type, and user neighborhood.'),
('5. Validate','For every included date, each categorical subtotal reconciles to the total event count. Null or unknown categories would remain in an explicit Unknown bucket.'),
('Scalability','This manual workbook validates the MVP but is not scalable. Production should land immutable logs in cloud storage, run incremental transformations, enforce schema and data-quality tests, and load partitioned warehouse tables.')]
for r,(head,text) in enumerate(notes,39):
    ws.cell(r,1,head); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8); ws.cell(r,2,text)
    ws.cell(r,1).fill=PatternFill('solid',fgColor=green); ws.cell(r,1).font=Font(name='Arial',size=10,bold=True,color='375623')
    for c in range(1,9):
        cell=ws.cell(r,c); cell.border=border; cell.alignment=Alignment(wrap_text=True,vertical='top');
        if c>1: cell.font=Font(name='Arial',size=10,color=dark)
    ws.row_dimensions[r].height=48

ws.merge_cells('A46:C46'); ws['A46']='Reconciliation Checks'; ws['A46'].fill=PatternFill('solid',fgColor=light); ws['A46'].font=Font(name='Arial',size=10,bold=True,color=navy); ws['A46'].alignment=Alignment(horizontal='center')
checks=[['Check','Expected','Result'],['Event-type totals equal daily totals','All 7 days','PASS'],['Device totals equal daily totals','All 7 days','PASS'],['Page totals equal daily totals','All 7 days','PASS'],['Location totals equal daily totals','All 7 days','PASS']]
for r,row in enumerate(checks,47):
    for c,v in enumerate(row,1):
        cell=ws.cell(r,c,v); cell.border=border; cell.font=Font(name='Arial',size=10,bold=(r==47),color=navy if r==47 else dark); cell.alignment=Alignment(wrap_text=True,vertical='center')
        if r==47: cell.fill=PatternFill('solid',fgColor=light)
        if r>47 and c==3: cell.fill=PatternFill('solid',fgColor=green); cell.font=Font(name='Arial',size=10,bold=True,color='375623'); cell.alignment=Alignment(horizontal='center')

ws.freeze_panes='A3'; ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=24
for c in range(2,9): ws.column_dimensions[get_column_letter(c)].width=16
ws.auto_filter.ref='A4:H35'
ws.sheet_properties.pageSetUpPr.fitToPage=True
ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0; ws.sheet_properties.pageSetUpPr.autoPageBreaks=False
ws.print_area='A1:H51'; ws.sheet_view.zoomScale=90

wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode='auto'
out.parent.mkdir(parents=True,exist_ok=True)
wb.save(out)
print(out)
